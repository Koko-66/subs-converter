# https://realpython.com/openpyxl-excel-spreadsheets-python/
from openpyxl.styles import (Border,
                             Font,
                             NamedStyle,
                             PatternFill,
                             Side,)


header = NamedStyle(name='header')
header.font = Font(bold=True, size=12)
header.border = Border(bottom=Side(border_style="medium"))
header.fill = PatternFill(
    fill_type="solid", start_color="00969696", end_color="00969696")


def format_header(ws):
    """Apply formatting to first row"""
    # for cell in ws[1]:
    #     cell.style = header
    row = ws.row_dimension[1, 1]
    row.style = header


# >>> from openpyxl.styles import PatternFill
# >>> from openpyxl.styles.differential import DifferentialStyle
# >>> from openpyxl.formatting.rule import Rule

# >>> red_background = PatternFill(fgColor="00FF0000")
# >>> diff_style = DifferentialStyle(fill=red_background)
# >>> rule = Rule(type="expression", dxf=diff_style)
# >>> rule.formula = ["$H1<3"]
# >>> sheet.conditional_formatting.add("A1:O100", rule)
# >>> workbook.save("sample_conditional_formatting.xlsx")